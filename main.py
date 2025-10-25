from fastapi import FastAPI, Depends, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy.sql import func
from typing import Optional, Dict, Any
from datetime import datetime, timedelta
from contextlib import asynccontextmanager
from io import BytesIO

import models
import schemas
import crud
from database import SessionLocal, engine, get_db
from auth import authenticate_user, create_access_token, get_current_user, require_role, create_superadmin, ACCESS_TOKEN_EXPIRE_MINUTES
from activity import get_activity_logs

# Force rebuild test 3 - Clean Pydantic v2 + Python 3.11.9 fix
print("Force rebuild test 3")

# Create database tables (including new orders and items tables)
models.Base.metadata.create_all(bind=engine)

# Ensure orders and items tables exist (backward compatibility)
try:
    from sqlalchemy import inspect
    inspector = inspect(engine)
    
    if 'orders' not in inspector.get_table_names():
        print("Creating orders table...")
        models.Order.__table__.create(bind=engine, checkfirst=True)
        print("✅ Orders table created successfully")
    
    if 'items' not in inspector.get_table_names():
        print("Creating items table...")
        models.Item.__table__.create(bind=engine, checkfirst=True)
        print("✅ Items table created successfully")
        
except Exception as e:
    print(f"Note: Table check: {str(e)}")

@asynccontextmanager
async def lifespan(app: FastAPI):
    # Startup
    db = SessionLocal()
    try:
        create_superadmin(db)
    finally:
        db.close()
    yield
    # Shutdown (if needed)

app = FastAPI(title="Inventory Management API", version="1.0.0", lifespan=lifespan)

# Add CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:3000", 
        "http://localhost:3001", 
        "https://ds-inventory-client.vercel.app",
        "http://localhost:3002",
        "https://ds-inventory-frontend.onrender.com",  # Render frontend URL
        "https://*.onrender.com",  # Allow all Render subdomains
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Authentication Endpoints
@app.post("/login", response_model=schemas.AuthResponse)
def login(user_credentials: schemas.UserLogin, db: Session = Depends(get_db)):
    """
    Authenticate user and return JWT token
    POST /login
    """
    try:
        user = authenticate_user(db, user_credentials.username, user_credentials.password)
        if not user:
            return schemas.AuthResponse(
                success=False,
                message="Invalid username or password",
                token=None
            )
        
        access_token_expires = timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
        access_token = create_access_token(
            data={"sub": user.username}, 
            expires_delta=access_token_expires
        )
        
        token = schemas.Token(
            access_token=access_token,
            token_type="bearer",
            user=user
        )
        
        return schemas.AuthResponse(
            success=True,
            message=f"Welcome back, {user.username}!",
            token=token
        )
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.post("/register", response_model=schemas.AuthResponse)
def register(
    user_data: Dict[str, Any], 
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_role(["superadmin", "admin"]))
):
    """
    Register a new user (superadmin and admin can do this)
    POST /register
    
    - Superadmin can create all roles (superadmin, admin, editor, viewer)
    - Admin can only create editor and viewer roles
    """
    try:
        print(f"Received user_data: {user_data}")  # Debug log
        
        # Validate required fields
        if not user_data.get('username'):
            return schemas.AuthResponse(
                success=False,
                message="Username is required",
                token=None
            )
        if not user_data.get('password'):
            return schemas.AuthResponse(
                success=False,
                message="Password is required",
                token=None
            )
        
        requested_role = user_data.get('role', 'viewer').lower()
        
        # Check if admin is trying to create admin or superadmin
        if current_user.role.value == "admin" and requested_role in ["admin", "superadmin"]:
            return schemas.AuthResponse(
                success=False,
                message="Admins can only create Editor or Viewer users",
                token=None
            )
        
        # Manually create UserCreate object
        try:
            user_create = schemas.UserCreate(
                username=user_data.get('username'),
                password=user_data.get('password'),
                role=requested_role
            )
        except ValueError as ve:
            print(f"Validation error creating UserCreate: {str(ve)}")
            return schemas.AuthResponse(
                success=False,
                message=f"Validation error: {str(ve)}",
                token=None
            )
        
        result = crud.create_user(db, user_create, current_user)
        
        if isinstance(result, dict) and "error" in result:
            return schemas.AuthResponse(
                success=False,
                message=result["error"],
                token=None
            )
        
        return schemas.AuthResponse(
            success=True,
            message=f"User {result.username} created successfully with role {result.role.value}",
            token=None
        )
    except ValueError as e:
        print(f"ValueError: {str(e)}")  # Debug log
        return schemas.AuthResponse(
            success=False,
            message=f"Validation error: {str(e)}",
            token=None
        )
    except Exception as e:
        print(f"Exception: {str(e)}")  # Debug log
        raise HTTPException(status_code=400, detail=str(e))

@app.get("/users", response_model=schemas.UsersResponse)
def get_users(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_role(["superadmin", "admin"]))
):
    """Get all users (superadmin and admin can access)"""
    try:
        users = crud.get_users(db)
        return schemas.UsersResponse(users=users)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.delete("/users/{user_id}", response_model=schemas.DeleteUserResponse)
def delete_user(
    user_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_role(["superadmin", "admin"]))
):
    """
    Delete a user (superadmin and admin can do this)
    - Admins cannot delete superadmin or admin users
    - Superadmins can delete any user except themselves
    """
    try:
        result = crud.delete_user(db, user_id, current_user)
        
        if isinstance(result, dict) and "error" in result:
            return schemas.DeleteUserResponse(
                success=False,
                message=result["error"]
            )
        
        if isinstance(result, dict) and "message" in result:
            return schemas.DeleteUserResponse(
                success=True,
                message=result["message"]
            )
        
        return schemas.DeleteUserResponse(
            success=False,
            message="Unexpected error occurred"
        )
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.get("/activity-logs", response_model=schemas.ActivityLogResponse)
def get_activity_history(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    """Get activity logs - all users can view"""
    try:
        logs = get_activity_logs(db)
        return schemas.ActivityLogResponse(logs=logs)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

# Item Management Endpoints
@app.post("/items/create", response_model=schemas.ItemCreateResponse)
def create_item(
    item_data: schemas.ItemCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_role(["superadmin", "admin", "editor"]))
):
    """
    Create a new item
    POST /items/create
    
    Accessible to: superadmin, admin, and editor
    """
    try:
        result = crud.create_item(db, item_data.item_name, current_user)
        
        if isinstance(result, dict) and "error" in result:
            return schemas.ItemCreateResponse(
                success=False,
                message=result["error"],
                item=None
            )
        
        return schemas.ItemCreateResponse(
            success=True,
            message=f"Item '{item_data.item_name}' created successfully",
            item=result
        )
    except Exception as e:
        print(f"Error in create_item endpoint: {str(e)}")
        raise HTTPException(status_code=400, detail=str(e))

@app.get("/items", response_model=schemas.ItemsListResponse)
def get_items(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    """
    Get all items
    GET /items
    
    Accessible to: all authenticated users
    """
    try:
        items = crud.get_all_items(db)
        return schemas.ItemsListResponse(items=items)
    except Exception as e:
        print(f"Error in get_items endpoint: {str(e)}")
        raise HTTPException(status_code=400, detail=str(e))

@app.post("/add", response_model=schemas.ProductResponse)
def add_product(
    request: schemas.AddProductRequest, 
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_role(["superadmin", "admin", "editor"]))
):
    """
    Add new stock to a product (requires add/edit permissions)
    POST /add
    """
    try:
        product = crud.add_product(db, request, current_user)
        return schemas.ProductResponse(
            success=True,
            message=f"Successfully added {request.quantity} units to {request.product_name} at ${request.unit_price} each (Total: ${request.quantity * request.unit_price})",
            product=product
        )
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

# SELL PRODUCT ENDPOINT - DISABLED
# The Sell Product feature has been removed from the project
# Orders module should be used instead for recording sales
# 
# @app.post("/sell", response_model=schemas.ProductResponse)
# def sell_product(
#     request: schemas.SellProductRequest, 
#     db: Session = Depends(get_db),
#     current_user: models.User = Depends(require_role(["superadmin", "admin", "editor"]))
# ):
#     """
#     Record a product sale (requires add/edit permissions)
#     POST /sell
#     """
#     try:
#         result = crud.sell_product(db, request, current_user)
#         
#         # Check if result contains error
#         if isinstance(result, dict) and "error" in result:
#             return schemas.ProductResponse(
#                 success=False,
#                 message=result["error"],
#                 product=None
#             )
#         
#         return schemas.ProductResponse(
#             success=True,
#             message=f"Successfully sold {request.quantity} units of {request.product_name} at ${request.unit_price} each (Total: ${request.quantity * request.unit_price})",
#             product=result
#         )
#     except Exception as e:
#         raise HTTPException(status_code=400, detail=str(e))

@app.get("/summary")
def get_summary(
    start: Optional[str] = None, 
    end: Optional[str] = None, 
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    """
    Get overview of all products or date range summary
    GET /summary
    GET /summary?start=2025-10-01&end=2025-10-31
    """
    try:
        if start and end:
            result = crud.get_date_range_summary(db, start, end)
            return schemas.DateRangeSummaryResponse(
                products=result["products"],
                total_added_qty_in_range=result["total_added_qty_in_range"],
                total_added_amount_in_range=result["total_added_amount_in_range"],
                total_sold_qty_in_range=result["total_sold_qty_in_range"],
                total_sold_amount_in_range=result["total_sold_amount_in_range"]
            )
        else:
            products = crud.get_summary(db)
            return schemas.SummaryResponse(products=products)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.get("/summary/enhanced", response_model=schemas.EnhancedSummaryResponse)
def get_enhanced_summary(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    """
    Get enhanced summary with financial analytics
    GET /summary/enhanced
    """
    try:
        enhanced_products = crud.get_enhanced_summary(db)
        return schemas.EnhancedSummaryResponse(products=enhanced_products)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.get("/products")
def get_products(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    """Get all product names for dropdown selection"""
    try:
        product_names = crud.get_all_products(db)
        return {"products": product_names}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.get("/products/details")
def get_products_with_details(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    """Get all products with full details (id, name, available_stock) for order creation"""
    try:
        products = crud.get_all_products_with_details(db)
        return {"products": products}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.get("/daily-history", response_model=schemas.TransactionHistoryResponse)
def get_daily_history(
    start: Optional[str] = None, 
    end: Optional[str] = None, 
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    """
    Get individual transaction history combining add and sell records
    GET /daily-history
    GET /daily-history?start=2025-10-01&end=2025-10-31
    
    Returns individual transaction records with:
    - date: Transaction date
    - product_name: Name of the product
    - transaction_type: "add" or "sell"
    - quantity: Quantity in transaction
    - unit_price: Price per unit
    - total_amount: Total amount for transaction
    
    Results are sorted by date descending (most recent first)
    """
    try:
        transactions = crud.get_transaction_history(db, start, end)
        return schemas.TransactionHistoryResponse(transactions=transactions)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.get("/database/view")
def view_database(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_role(["superadmin"]))
):
    """
    View all database data in a readable format
    GET /database/view
    """
    try:
        # Get all products
        products = db.query(models.Product).all()
        
        # Get all add history
        add_history = db.query(models.AddHistory).all()
        
        # Get all sell history  
        sell_history = db.query(models.SellHistory).all()
        
        return {
            "products": [
                {
                    "id": p.id,
                    "name": p.name,
                    "total_added_qty": p.total_added_qty,
                    "total_added_amount": str(p.total_added_amount),
                    "total_sold_qty": p.total_sold_qty,
                    "total_sold_amount": str(p.total_sold_amount),
                    "available_stock": p.available_stock
                } for p in products
            ],
            "add_history": [
                {
                    "id": a.id,
                    "product_id": a.product_id,
                    "product_name": db.query(models.Product).filter(models.Product.id == a.product_id).first().name,
                    "quantity": a.quantity,
                    "unit_price": str(a.unit_price),
                    "total_amount": str(a.total_amount),
                    "date": a.date
                } for a in add_history
            ],
            "sell_history": [
                {
                    "id": s.id,
                    "product_id": s.product_id,
                    "product_name": db.query(models.Product).filter(models.Product.id == s.product_id).first().name,
                    "quantity": s.quantity,
                    "unit_price": str(s.unit_price),
                    "total_amount": str(s.total_amount),
                    "date": s.date
                } for s in sell_history
            ],
            "summary": {
                "total_products": len(products),
                "total_transactions": len(add_history) + len(sell_history),
                "total_add_transactions": len(add_history),
                "total_sell_transactions": len(sell_history)
            }
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.delete("/history/add/{add_id}", response_model=schemas.DeleteResponse)
def delete_add_history(
    add_id: int, 
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_role(["superadmin", "admin"]))
):
    """
    Delete an add history record and update product totals
    DELETE /history/add/{id}
    """
    try:
        result = crud.delete_add_history(db, add_id, current_user)
        
        # Check if result contains error
        if isinstance(result, dict) and "error" in result:
            return schemas.DeleteResponse(
                success=False,
                message=result["error"],
                updated_product=None
            )
        
        return schemas.DeleteResponse(
            success=True,
            message=f"Successfully deleted add history record (ID: {add_id}). Product totals updated.",
            updated_product=result
        )
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.delete("/history/sell/{sell_id}", response_model=schemas.DeleteResponse)
def delete_sell_history(
    sell_id: int, 
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_role(["superadmin", "admin"]))
):
    """
    Delete a sell history record and update product totals
    DELETE /history/sell/{id}
    """
    try:
        result = crud.delete_sell_history(db, sell_id, current_user)
        
        # Check if result contains error
        if isinstance(result, dict) and "error" in result:
            return schemas.DeleteResponse(
                success=False,
                message=result["error"],
                updated_product=None
            )
        
        return schemas.DeleteResponse(
            success=True,
            message=f"Successfully deleted sell history record (ID: {sell_id}). Product totals updated.",
            updated_product=result
        )
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

# User Profile Endpoints
@app.get("/user/me", response_model=schemas.UserMeResponse)
def get_current_user_info(
    current_user: models.User = Depends(get_current_user)
):
    """
    Get current user's profile information
    GET /user/me
    """
    try:
        return schemas.UserMeResponse(
            id=current_user.id,
            username=current_user.username,
            name=current_user.name,
            role=current_user.role.value if hasattr(current_user.role, 'value') else str(current_user.role),
            created_at=current_user.created_at
        )
    except Exception as e:
        print(f"Error fetching user info: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to fetch user information")

@app.put("/user/update_profile", response_model=schemas.ProfileUpdateResponse)
def update_user_profile(
    profile_data: schemas.ProfileUpdateRequest,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    """
    Update current user's profile
    PUT /user/update_profile
    """
    try:
        # Get the user from database
        user = db.query(models.User).filter(models.User.id == current_user.id).first()
        if not user:
            raise HTTPException(status_code=404, detail="User not found")
        
        # Update name if provided
        if profile_data.name is not None:
            user.name = profile_data.name.strip() if profile_data.name.strip() else None
        
        # Update username if provided and different
        if profile_data.username is not None and profile_data.username != user.username:
            # Check if username already exists
            existing_user = db.query(models.User).filter(
                models.User.username == profile_data.username,
                models.User.id != user.id
            ).first()
            
            if existing_user:
                raise HTTPException(status_code=400, detail="Username already exists")
            
            user.username = profile_data.username
        
        # Update password if provided
        if profile_data.password is not None:
            from auth import get_password_hash
            user.password_hash = get_password_hash(profile_data.password)
        
        # Commit changes to database
        db.commit()
        db.refresh(user)
        
        # Log activity for audit trail
        try:
            from activity import log_activity
            log_activity(
                db,
                user,
                "Update Profile",
                f"user {user.username}",
                f"User updated their profile"
            )
        except Exception as log_error:
            print(f"Warning: Failed to log activity: {str(log_error)}")
        
        return schemas.ProfileUpdateResponse(message="Profile updated successfully")
        
    except HTTPException:
        # Re-raise HTTP exceptions (like 400, 404)
        raise
    except Exception as e:
        # Rollback any database changes
        db.rollback()
        print(f"Error updating user profile: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to update profile")

@app.get("/")
def root():
    return {"message": "Inventory Management API is running!", "status": "healthy", "version": "1.0.0"}

@app.get("/health")
def health_check():
    """Health check endpoint for monitoring"""
    return {
        "status": "healthy",
        "message": "API is running",
        "timestamp": datetime.now().isoformat(),
        "version": "1.0.0"
    }

# Order Management Endpoints
@app.post("/orders/create", response_model=schemas.CreateOrderResponse)
def create_order(
    request: schemas.CreateOrderRequest,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_role(["superadmin", "admin", "editor"]))
):
    """
    Create a new order with customer details
    POST /orders/create
    
    Requires: superadmin, admin, or editor role
    """
    try:
        result = crud.create_order(db, request, current_user)
        
        # Check if result contains error
        if isinstance(result, dict) and "error" in result:
            return schemas.CreateOrderResponse(
                success=False,
                message=result["error"],
                order=None
            )
        
        return schemas.CreateOrderResponse(
            success=True,
            message=f"Order created successfully! Sold {request.quantity_sold} units of {request.product_name} to {request.customer_name or 'Customer'}",
            order=result
        )
    except Exception as e:
        print(f"Error in create_order endpoint: {str(e)}")
        raise HTTPException(status_code=400, detail=str(e))

@app.get("/orders", response_model=schemas.OrdersResponse)
def get_orders(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    product_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_role(["superadmin", "admin", "editor"]))
):
    """
    Get all orders with optional filters
    GET /orders
    GET /orders?start_date=2025-10-01&end_date=2025-10-31
    GET /orders?product_id=1
    
    Accessible to: superadmin, admin, and editor
    """
    try:
        orders = crud.get_orders(db, start_date, end_date, product_id)
        return schemas.OrdersResponse(orders=orders)
    except Exception as e:
        print(f"Error in get_orders endpoint: {str(e)}")
        raise HTTPException(status_code=400, detail=str(e))

@app.get("/orders/export")
def export_orders_to_excel(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_role(["superadmin", "admin"]))
):
    """
    Export all orders to Excel file
    GET /orders/export
    
    Only accessible to: superadmin and admin
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Get all orders
        orders = crud.get_orders(db)
        
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Orders"
        
        # Define headers
        headers = [
            "Order ID",
            "Product",
            "Quantity",
            "Customer Name",
            "Phone",
            "Address",
            "Total Amount",
            "Sale Date",
            "Created By"
        ]
        
        # Style for headers
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Write data
        for row_num, order in enumerate(orders, 2):
            ws.cell(row=row_num, column=1, value=order.id)
            ws.cell(row=row_num, column=2, value=order.product_name)
            ws.cell(row=row_num, column=3, value=order.quantity_sold)
            ws.cell(row=row_num, column=4, value=order.customer_name or "N/A")
            ws.cell(row=row_num, column=5, value=order.customer_phone or "N/A")
            ws.cell(row=row_num, column=6, value=order.customer_address or "N/A")
            ws.cell(row=row_num, column=7, value=f"${float(order.total_amount):.2f}")
            ws.cell(row=row_num, column=8, value=str(order.sale_date))
            ws.cell(row=row_num, column=9, value=order.created_by)
        
        # Adjust column widths
        column_widths = [10, 20, 10, 20, 15, 30, 15, 20, 15]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col_num)].width = width
        
        # Save to BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"orders_export_{timestamp}.xlsx"
        
        # Return as streaming response
        return StreamingResponse(
            excel_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        print(f"Error exporting orders to Excel: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to export orders: {str(e)}")

if __name__ == "__main__":
    import uvicorn
    import os
    port = int(os.getenv("PORT", 8000))
    uvicorn.run(app, host="0.0.0.0", port=port)
